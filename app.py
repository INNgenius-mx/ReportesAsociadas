"""
Reporte de Asociados — Procesador de PDF (Betterware)
======================================================
App local (Streamlit). Sube el PDF "Actividad de Mis Asociadas", extrae los datos
(código, teléfono, pedidos y TOTAL DE COMPRAS DEL CATÁLOGO ACTUAL) del texto directo,
y usa IA (Claude o Grok) SOLO para reconstruir los nombres correctos. Genera Excel
y PDF formales, ordenados por compras del catálogo actual (desc) y desempate por pedidos.

La API key NO se guarda en el código: se pega en la app o se lee de variable de entorno.
"""

import io
import os
import re
import json
import streamlit as st


# ============================================================================
# CARGA DE API KEYS DESDE ARCHIVO .env (si existe, junto a app.py)
# El .env NO es parte del código; lo creas tú y guardas ahí tus keys.
# Formato del .env:
#   ANTHROPIC_API_KEY=sk-ant-...
#   XAI_API_KEY=xai-...
# ============================================================================
def cargar_env():
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(ruta):
        return
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea or linea.startswith("#") or "=" not in linea:
                    continue
                clave, valor = linea.split("=", 1)
                clave = clave.strip()
                valor = valor.strip().strip('"').strip("'")
                if clave and valor and clave not in os.environ:
                    os.environ[clave] = valor
    except Exception:
        pass


def cargar_streamlit_secrets():
    """Carga API keys desde Streamlit Secrets cuando la app está en la nube."""
    try:
        for clave in ("ANTHROPIC_API_KEY", "GEMINI_API_KEY", "XAI_API_KEY", "CLAVE_APP"):
            if clave not in os.environ:
                valor = st.secrets.get(clave, "")
                if valor:
                    os.environ[clave] = str(valor)
    except Exception:
        pass


cargar_env()
cargar_streamlit_secrets()

# ============================================================================
# UTILIDADES
# ============================================================================
def title_es(nombre):
    minus = {"de", "la", "las", "los", "del", "y", "da", "e"}
    palabras = nombre.strip().lower().split()
    out = []
    for i, p in enumerate(palabras):
        out.append(p if (p in minus and i != 0) else (p[:1].upper() + p[1:]))
    return " ".join(out)


def limpiar_dinero(txt):
    if txt is None:
        return 0
    s = re.sub(r"[^\d.]", "", str(txt))
    if s == "":
        return 0
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


# ============================================================================
# EXTRACCIÓN DE DATOS (texto directo, sin OCR)
# Los datos numéricos salen perfectos. El nombre crudo es de respaldo por si no se usa IA.
#
# IMPORTANTE: cada asociada SIEMPRE tiene una línea con "Sem:NN / AAAA" seguida de
# importes semanales variables, total de pedidos y total de compras. Esa línea es
# el ANCLA fiable de cada fila. El teléfono PUEDE NO EXISTIR, y el código a veces
# queda en una línea aparte (cuando el nombre es largo y se parte en varias líneas).
# Por eso ya NO se exige teléfono ni que el código esté pegado al inicio de la línea:
# así no se salta ninguna asociada.
# ============================================================================

# Patrón base para encontrar la parte financiera de cada asociada.
#
# El PDF puede cambiar de catálogo y traer 4, 5 o más semanas visibles:
#   Ejemplo 1: S23 S24 S25 S26
#   Ejemplo 2: S27 S28 S29 S30 S31
#
# Por eso ya NO contamos semanas fijas. Primero encontramos "Sem:NN / AAAA" y luego
# leemos los últimos campos de la fila desde la derecha:
#   ... semanas ... TotalPedidosActual TotalComprasActual Referidas SaldoPuntos
#   TotalPedidosAnterior TotalComprasAnterior
#
# Esto evita jalar el Total Compras del catálogo anterior y soporta PDFs con diferente
# cantidad de columnas semanales.
SEMANA_FILA = re.compile(r"Sem:\s*\d+\s*/\s*\d{4}")
TOKEN_NUMERICO = re.compile(r"\$[\d,]+|\d[\d,]*")

# Línea que contiene ÚNICAMENTE un código (con posibles símbolos * + ~ - al inicio).
SOLO_CODIGO = re.compile(r"^[\*\+\~\-\s]*(\d{6,8})\s*$")


def _nombre_fallback(prefijo):
    """Arma un nombre de respaldo (sin IA) con el fragmento alfabético del prefijo."""
    frag = re.sub(r"\(cid:[^)]*\)", " ", prefijo)        # quita basura (cid:..)
    frag = re.sub(r"\d+", " ", frag)                      # quita números (código/teléfono)
    frag = re.sub(r"[\*\+\~\$/]", " ", frag)              # quita símbolos
    frag = re.sub(r"\bSem:\b", " ", frag)
    frag = re.sub(r"\b(PLUS|BASE|Inactivo)\b", " ", frag, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", frag).strip()




# ----------------------------
# NOMBRES POR COORDENADAS
# ----------------------------
# En estos PDFs, pdfplumber extrae los números bien por línea, pero los nombres
# pueden quedar arriba/abajo de la línea que contiene "Sem:NN / AAAA". Por eso
# también leemos la posición visual de las palabras dentro de la columna "Nombre".
CID_RE = re.compile(r"\(cid:(\d+)\)")

# Mapa para palabras que vienen dañadas por la fuente del PDF.
# Ejemplos: 5RGUtJXH] -> Rodríguez, 3(f$ -> Peña, *XWLpUUH] -> Gutiérrez.
CIFRA_NOMBRE = {
    "5": "R", "R": "o", "G": "d", "U": "r", "H": "e", "X": "u",
    "J": "g", "]": "z", "&": "C", "$": "A", "(": "E", "3": "P",
    "L": "i", "O": "l", "F": "c", "D": "a", "Q": "n", "V": "s",
    "-": "J", "W": "t", "p": "é", "i": "á", "t": "í", "y": "ó",
    "x": "ñ", ")": "F", "0": "M", "2": "O", "*": "G", "f": "Ñ",
    "1": "N", "8": "ú", "=": "z", "/": "L",
}

PALABRAS_NO_NOMBRE = {
    "base", "plus", "inactivo", "distribuido", "con", "r", "ctbt", "me",
    "asociada", "asociado", "nombre", "fecha", "ingreso", "total", "pedidos",
    "compras", "referidas", "saldo", "puntos", "nivel", "cierre",
}


def _cid_a_texto(txt):
    """Convierte piezas tipo (cid:42) a su carácter base."""
    return CID_RE.sub(lambda m: chr(int(m.group(1))), str(txt))


def _descifrar_si_aplica(txt_original):
    """Descifra SOLO palabras que vienen como cid o con símbolos de la fuente dañada."""
    txt = _cid_a_texto(txt_original)
    # No tocar palabras normales como Estrada, Garcia, Elizabeth, etc.
    # Solo descifrar si el token venía como cid o trae símbolos/dígitos del código dañado.
    if "(cid:" in str(txt_original) or re.search(r"[0-9\]\&\$\(\)\*\-]", txt):
        return "".join(CIFRA_NOMBRE.get(ch, ch) for ch in txt)
    return txt


def _limpiar_nombre_desde_partes(partes):
    """Limpia y une palabras de nombre detectadas visualmente."""
    palabras = []
    for parte in partes:
        txt = _descifrar_si_aplica(parte)
        # Mantener espacios cuando una palabra dañada contiene dos palabras: Molina Peña.
        txt = re.sub(r"[^A-Za-zÁÉÍÓÚÜáéíóúüÑñ]+", " ", txt)
        for p in txt.split():
            if not p:
                continue
            if p.lower() in PALABRAS_NO_NOMBRE:
                continue
            palabras.append(p)

    nombre = " ".join(palabras)
    nombre = re.sub(r"\bGuadarram\s+A\b", "Guadarrama", nombre, flags=re.IGNORECASE)
    nombre = re.sub(r"\bGuillermin\s+A\b", "Guillermina", nombre, flags=re.IGNORECASE)
    nombre = re.sub(r"\bVelzquez\b", "Velázquez", nombre, flags=re.IGNORECASE)
    nombre = re.sub(r"\bVernica\b", "Verónica", nombre, flags=re.IGNORECASE)
    nombre = re.sub(r"\bLucia\b", "Lucía", nombre, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", nombre).strip()


def _extraer_nombres_por_coordenadas(page):
    """Devuelve {codigo: nombre} leyendo la columna de nombres por posición visual.

    Esto evita que salgan nombres como '(revisar nombre)' cuando el nombre está
    partido en varias líneas alrededor de la fila financiera.
    """
    nombres = {}
    try:
        words = page.extract_words(
            x_tolerance=2,
            y_tolerance=3,
            keep_blank_chars=False,
            use_text_flow=False,
        )
    except Exception:
        return nombres

    # Cada fila real tiene una palabra tipo Sem:39, Sem:27, etc.
    sems = [w for w in words if _cid_a_texto(w.get("text", "")).startswith("Sem:")]
    sems.sort(key=lambda w: w["top"])
    if not sems:
        return nombres

    for i, sem in enumerate(sems):
        prev_top = sems[i - 1]["top"] if i > 0 else sem["top"] - 60
        next_top = sems[i + 1]["top"] if i + 1 < len(sems) else sem["top"] + 60
        top_b = (prev_top + sem["top"]) / 2
        bot_b = (sem["top"] + next_top) / 2

        fila_words = [w for w in words if top_b - 1 <= w["top"] < bot_b + 1]

        # Código: 6 a 8 dígitos, normalmente a la izquierda de la columna de nombre.
        codigos = []
        for w in fila_words:
            raw = _cid_a_texto(w.get("text", ""))
            dig = re.sub(r"\D", "", raw)
            if 6 <= len(dig) <= 8 and w.get("x0", 9999) < 200:
                codigos.append(dig)
        if not codigos:
            continue
        codigo = codigos[0]

        # Nombre: columna visual entre código y teléfono.
        # En este formato Betterware suele caer entre x≈198 y x≈276.
        partes_nombre = []
        for w in fila_words:
            x0 = w.get("x0", 0)
            raw = _cid_a_texto(w.get("text", ""))
            if 198 <= x0 <= 276:
                if raw.startswith("Sem:"):
                    continue
                if re.fullmatch(r"\$?[\d,]+", raw):
                    continue
                partes_nombre.append(w.get("text", ""))

        nombre = _limpiar_nombre_desde_partes(partes_nombre)
        if nombre:
            nombres[codigo] = title_es(nombre)

    return nombres


def extraer(pdf_bytes):
    """Devuelve (filas, texto_crudo). filas con datos correctos + nombre crudo de respaldo.

    Procesa página por página. Para cada línea con el ancla financiera crea una fila;
    el teléfono es opcional y, si el código no está en esa línea, lo recupera de una
    línea vecina que contenga solo el código.
    """
    import pdfplumber
    paginas = []
    nombres_por_codigo = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            paginas.append((page.extract_text() or "").split("\n"))
            nombres_por_codigo.update(_extraer_nombres_por_coordenadas(page))
    texto_crudo = "\n".join("\n".join(p) for p in paginas)

    filas = []
    for lineas in paginas:
        codigos_usados = set()
        for idx, ln in enumerate(lineas):
            m = SEMANA_FILA.search(ln)
            if not m:
                continue

            # Tomamos la información DESPUÉS de "Sem:NN / AAAA".
            # Ejemplo con 5 semanas:
            # $2,800 $0 $0 $0 $0 1 $2,800 0 66,115 4 $9,453
            # Los últimos 6 tokens siempre son:
            # pedidos_actuales, compras_actuales, referidas, saldo_puntos,
            # pedidos_anteriores, compras_anteriores.
            resto = ln[m.end():]
            tokens = TOKEN_NUMERICO.findall(resto)

            if len(tokens) < 6:
                continue

            try:
                pedidos = int(tokens[-6].replace(",", ""))
                compras_actuales = limpiar_dinero(tokens[-5])
                compras_anteriores = limpiar_dinero(tokens[-1])
                pedidos_anteriores = int(tokens[-2].replace(",", ""))
            except Exception:
                continue

            prefijo = ln[:m.start()]
            # En el prefijo solo hay números de código (6-8 dígitos) y/o teléfono (>=9).
            nums = re.findall(r"\d{6,}", prefijo)
            codigo = next((n for n in nums if 6 <= len(n) <= 8), None)
            telefono = next((n for n in nums if len(n) >= 9), "")

            # Si el código no venía en la línea (nombre largo partido), buscarlo en una
            # línea cercana. Acepta tanto una línea con SOLO el código como una que
            # empiece con el código seguido de un fragmento del nombre
            # (p. ej. "14503899 Lopez Fuentes"). Nunca toma el código de otra fila de datos.
            if codigo is None:
                for off in (1, -1, 2, -2, 3, -3):
                    j = idx + off
                    if not (0 <= j < len(lineas)):
                        continue
                    cand = lineas[j]
                    if "Sem:" in cand:          # esa línea es otra fila de datos: no tomar su código
                        continue
                    mc = re.match(r"^[\*\+\~\-\s]*(\d{6,8})\b", cand)
                    if mc and mc.group(1) not in codigos_usados:
                        codigo = mc.group(1)
                        break
            if codigo:
                codigos_usados.add(codigo)

            nombre = ""
            if codigo:
                nombre = nombres_por_codigo.get(codigo, "")
            if not nombre:
                nombre = _nombre_fallback(prefijo)
                nombre = title_es(nombre) if nombre else "(revisar nombre)"

            filas.append({
                "codigo": codigo or "(revisar codigo)",
                "nombre": nombre,
                "telefono": telefono,
                "pedidos": pedidos,
                # Se conserva la llave "ingresos" para no romper el resto del código,
                # pero el valor real es Total Compras del Catálogo Actual.
                "ingresos": compras_actuales,
                "compras_catalogo_actual": compras_actuales,
                "pedidos_catalogo_actual": pedidos,
                "compras_catalogo_anterior": compras_anteriores,
                "pedidos_catalogo_anterior": pedidos_anteriores,
            })
    return filas, texto_crudo


def extraer_total_oficial_catalogo_actual(texto_crudo):
    """Intenta leer el total general oficial del PDF original.

    En los PDFs Betterware normalmente aparece una línea final así:
    Total $24,366 $0 $0 $0 $0 21 $24,366 0 81 $91,655

    El primer importe después de "Total" y el importe después de Total Pedidos
    corresponden a compras del Catálogo Actual. Si no se encuentra, devuelve None.
    """
    candidatos = []
    for ln in texto_crudo.splitlines():
        if not re.search(r"\bTotal\b", ln, flags=re.IGNORECASE):
            continue
        if "Sem:" in ln:
            continue
        tokens = TOKEN_NUMERICO.findall(ln)
        # Línea total con semanas + pedidos actual + compras actual + anterior.
        if len(tokens) >= 8:
            # Ejemplo con 5 semanas: $24,366 $0 $0 $0 $0 21 $24,366 0 81 $91,655
            # Los últimos 6 tokens son: pedidos_actual, compras_actual, referidas,
            # saldo_puntos si existe, pedidos_anterior, compras_anterior. En la línea
            # Total, puede faltar saldo_puntos, por eso preferimos el último importe
            # antes del bloque de catálogo anterior: tokens[-4] en este formato.
            try:
                compras_actual = limpiar_dinero(tokens[-4])
                pedidos_actual = int(str(tokens[-5]).replace(",", ""))
                candidatos.append((pedidos_actual, compras_actual, ln.strip()))
            except Exception:
                pass
    if not candidatos:
        return None
    # Tomamos el último candidato porque el total general suele estar en la última página.
    return candidatos[-1]


# ============================================================================
# LIMPIEZA DE NOMBRES CON IA (Claude o Grok)
# ============================================================================
def limpiar_nombres_con_ia(texto_crudo, filas, proveedor, api_key):
    codigos = [f["codigo"] for f in filas]
    prompt = (
        "Eres un experto extrayendo nombres de un reporte de Betterware ('Actividad de Mis Asociadas'). "
        "En el texto, cada asociada tiene un CÓDIGO numérico y, cerca de él, su nombre completo, que suele "
        "venir PARTIDO en varias líneas (por ejemplo 'MARIA', 'DEL', 'ROSARIO', 'CHAVEZ', 'AVILES' en "
        "líneas distintas). Une esos fragmentos para formar el nombre completo de cada código.\n\n"
        "INSTRUCCIONES:\n"
        "- Incluye SIEMPRE el nombre COMPLETO: nombre(s) y TODOS los apellidos. No omitas ningún apellido, "
        "aunque venga en una línea aparte o rodeado de números o símbolos.\n"
        "- Escribe los nombres con sus acentos y la Ñ correctos en español (María, Núñez, Peña, Fernández, "
        "Rodríguez, Gutiérrez, Garduño, Castañeda, etc.).\n"
        "- TEXTO CIFRADO: por la fuente del PDF, ALGUNOS nombres vienen 'revueltos' (mezcla de mayúsculas, "
        "dígitos y símbolos, p. ej. '5RGUtJXH]', '-XiUH]', '3LxD', '&DxHGR', '9HOi]TXH]'). Eso es un nombre. "
        "Descífralo con esta clave (carácter mostrado -> letra real): "
        "5->R R->o G->d U->r H->e X->u J->g ]->z &->C $->A (->E 3->P L->i O->l F->c D->a Q->n V->s -->J ; "
        "y para acentos/ñ: i->á t->í y->ó x->ñ f->Ñ . "
        "Ejemplos: '5RGUtJXH]'->Rodríguez, '-XiUH]'->Juárez, '3LxD'->Piña, '&DxHGR'->Cañedo, "
        "'9HOi]TXH]'->Velázquez. Aplica esto SOLO a las palabras revueltas de nombres; NUNCA a números, "
        "teléfonos ni a 'Sem:'.\n"
        "- Si algún carácter del texto está dañado o ilegible, dedúcelo según cómo se escribe normalmente "
        "ese nombre en español. No cambies un nombre por otro distinto (por ejemplo, NO conviertas Piña en "
        "Lira, ni Juárez en Núñez).\n"
        "- Formato de salida: cada nombre como 'Nombre Apellido Apellido' (primera letra de cada palabra en "
        "mayúscula; partículas como de, del, la, los en minúscula).\n\n"
        "Responde ÚNICAMENTE con un JSON válido (sin texto adicional, sin bloques de código): un objeto "
        "donde la llave es el código (string) y el valor es el nombre completo. "
        "Debes incluir los " + str(len(codigos)) + " códigos. Códigos: " + ", ".join(codigos) +
        "\n\nTEXTO CRUDO:\n" + texto_crudo
    )
    if proveedor == "Claude":
        contenido = _llamar_claude(prompt, api_key)
    elif proveedor == "Grok":
        contenido = _llamar_grok(prompt, api_key)
    else:
        contenido = _llamar_gemini(prompt, api_key)

    contenido = contenido.strip()
    contenido = re.sub(r"^```(?:json)?", "", contenido).strip()
    contenido = re.sub(r"```$", "", contenido).strip()
    ini, fin = contenido.find("{"), contenido.rfind("}")
    if ini != -1 and fin != -1:
        contenido = contenido[ini:fin + 1]

    # Intento 1: parsear el JSON completo
    mapa = {}
    try:
        mapa = json.loads(contenido)
    except json.JSONDecodeError:
        # Intento 2 (respaldo): si vino cortado, rescatar pares "codigo": "nombre" con regex
        for m in re.finditer(r'"(\d{6,8})"\s*:\s*"([^"]+)"', contenido):
            mapa[m.group(1)] = m.group(2)
        if not mapa:
            raise RuntimeError("La IA devolvió un formato que no se pudo leer.")

    for f in filas:
        val = mapa.get(f["codigo"])
        if val and str(val).strip():
            f["nombre"] = title_es(str(val).strip())
    return filas


def _llamar_claude(prompt, api_key):
    import urllib.request, urllib.error
    body = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 4096,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=body,
        headers={"content-type": "application/json", "x-api-key": api_key,
                 "anthropic-version": "2023-06-01"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detalle = e.read().decode("utf-8", "ignore")
        raise RuntimeError(f"HTTP {e.code} de Anthropic: {detalle}")
    return "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")


def _llamar_grok(prompt, api_key):
    import urllib.request, urllib.error
    body = json.dumps({
        "model": "grok-4.3",
        "messages": [{"role": "user", "content": prompt}],
        "max_completion_tokens": 4096,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.x.ai/v1/chat/completions", data=body,
        headers={"content-type": "application/json", "authorization": "Bearer " + api_key},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detalle = e.read().decode("utf-8", "ignore")
        raise RuntimeError(f"HTTP {e.code} de xAI/Grok: {detalle}")
    return data["choices"][0]["message"]["content"]


def _llamar_gemini(prompt, api_key):
    import urllib.request, urllib.error, socket
    modelo = "gemini-2.5-flash"
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{modelo}:generateContent"
    body = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 65536,
                              "responseMimeType": "application/json",
                              "thinkingConfig": {"thinkingBudget": 0}},
    }).encode("utf-8")
    req = urllib.request.Request(
        url, data=body,
        headers={"content-type": "application/json", "x-goog-api-key": api_key},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detalle = e.read().decode("utf-8", "ignore")
        raise RuntimeError(f"HTTP {e.code} de Gemini: {detalle}")
    except (socket.timeout, TimeoutError):
        raise RuntimeError("Gemini tardó demasiado (timeout). Intenta de nuevo.")
    # La respuesta viene en candidates[0].content.parts[*].text
    try:
        partes = data["candidates"][0]["content"]["parts"]
        return "".join(p.get("text", "") for p in partes)
    except (KeyError, IndexError):
        raise RuntimeError(f"Respuesta inesperada de Gemini: {json.dumps(data)[:500]}")


# ============================================================================
# EXCEL
# ============================================================================
def generar_excel(filas, titulo_reporte):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import DataBarRule

    MORADO, MORADO_OSC, NARANJA, BLANCO, LILA = "702080", "4A1560", "F25C00", "FFFFFF", "F3ECF7"
    thin = Side(style="thin", color="D8CCE2")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Nombre del Asociado", "Pedidos", "Compras Catálogo Actual"]

    def construir(ws, subtitulo, registros):
        """Llena una hoja con la tabla de marca (título, logo, encabezado, datos, total, barras)."""
        ws.sheet_view.showGridLines = False

        # Título
        ws.merge_cells("A1:C1")
        c = ws["A1"]; c.value = titulo_reporte
        c.font = Font(name="Arial", size=16, bold=True, color=BLANCO)
        c.fill = PatternFill("solid", fgColor=MORADO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34
        try:
            from openpyxl.drawing.image import Image as XLImage
            lp = _ruta_junto_app("logo.png")
            if os.path.exists(lp):
                img = XLImage(lp); img.height = 40; img.width = int(40 * 271 / 301)
                ws.add_image(img, "A1")
        except Exception:
            pass

        # Subtítulo naranja (marca + descripción de la hoja)
        ws.merge_cells("A2:C2")
        s = ws["A2"]; s.value = f"INNquietus · {subtitulo}"
        s.font = Font(name="Arial", size=11, bold=True, color=BLANCO)
        s.fill = PatternFill("solid", fgColor=NARANJA)
        s.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # Encabezados
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = Font(name="Arial", size=11, bold=True, color=BLANCO)
            cell.fill = PatternFill("solid", fgColor=MORADO_OSC)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
        ws.row_dimensions[3].height = 28

        # Datos
        fila = 4
        for idx, r in enumerate(registros):
            relleno = LILA if idx % 2 else BLANCO
            vals = [r["nombre"], r["pedidos"], r["ingresos"]]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=fila, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = PatternFill("solid", fgColor=relleno)
                cell.border = borde
                if col == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col == 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = '"$"#,##0'
            fila += 1

        # Total
        ws.cell(row=fila, column=1, value="TOTAL")
        for col in (1, 2, 3):
            cell = ws.cell(row=fila, column=col)
            cell.font = Font(name="Arial", size=11, bold=True, color=BLANCO)
            cell.fill = PatternFill("solid", fgColor=MORADO)
            cell.border = borde
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if registros:
            ws.cell(row=fila, column=2, value=f"=SUM(B4:B{fila-1})").alignment = Alignment(horizontal="center", vertical="center")
            e = ws.cell(row=fila, column=3, value=f"=SUM(C4:C{fila-1})")
        else:
            ws.cell(row=fila, column=2, value=0).alignment = Alignment(horizontal="center", vertical="center")
            e = ws.cell(row=fila, column=3, value=0)
        e.alignment = Alignment(horizontal="right", vertical="center"); e.number_format = '"$"#,##0'

        # Barras de datos moradas en la columna de compras (más color)
        if registros:
            regla = DataBarRule(start_type="min", end_type="max", color="B583C9", showValue=True)
            ws.conditional_formatting.add(f"C4:C{fila-1}", regla)

        for col, w in zip("ABC", (44, 14, 26)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A4"

    # Separar: activas = 1 o más pedidos; inactivas = 0 pedidos
    activos = [r for r in filas if r["pedidos"] >= 1]
    inactivos = [r for r in filas if r["pedidos"] == 0]

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Reporte"
    construir(ws1, "Todos · Ordenado por compras del catálogo actual", filas)                     # tabla original completa
    construir(wb.create_sheet("Activos"), "Activos · 1 o más pedidos", activos)
    construir(wb.create_sheet("Inactivos"), "Inactivos · 0 pedidos", inactivos)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ============================================================================
# PDF
# ============================================================================
def generar_pdf(filas, titulo_reporte):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=1.3*cm, bottomMargin=1.5*cm,
                            leftMargin=1.2*cm, rightMargin=1.2*cm)
    estilos = getSampleStyleSheet()
    marca_style = ParagraphStyle("marca", parent=estilos["Normal"], fontSize=11, alignment=TA_CENTER,
                                 textColor=colors.HexColor("#F25C00"), spaceAfter=2, leading=13)
    titulo_style = ParagraphStyle("titulo", parent=estilos["Title"], textColor=colors.HexColor("#702080"),
                                  fontSize=18, alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=estilos["Normal"], fontSize=9, alignment=TA_CENTER,
                               textColor=colors.HexColor("#404040"), spaceAfter=12)

    elementos = []
    logo_path = _ruta_junto_app("logo.png")
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=1.7*cm, height=1.7*cm*301/271)  # mantiene proporción
            logo.hAlign = "CENTER"
            elementos += [logo, Spacer(1, 4)]
        except Exception:
            pass
    elementos += [Paragraph("<b>INNquietus</b>", marca_style),
                  Paragraph(titulo_reporte, titulo_style),
                  Paragraph("Catálogo Actual · Ordenado por compras del catálogo actual", sub_style), Spacer(1, 6)]

    def banda(texto, color_hex):
        """Banner de color de ancho completo para encabezar cada sección."""
        b = Table([[texto]], colWidths=[17*cm])
        b.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(color_hex)),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        return b

    def construir_tabla(registros):
        header_style = ParagraphStyle(
            "header_tabla",
            parent=estilos["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=8,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        data = [[
            Paragraph("Nombre del Asociado", header_style),
            Paragraph("Pedidos", header_style),
            Paragraph("Compras Catálogo Actual", header_style),
        ]]
        tot_p = tot_i = 0
        for r in registros:
            data.append([r["nombre"], str(r["pedidos"]), f"${r['ingresos']:,}"])
            tot_p += r["pedidos"]; tot_i += r["ingresos"]
        data.append(["TOTAL", str(tot_p), f"${tot_i:,}"])
        t = Table(data, colWidths=[9.4*cm, 2.6*cm, 5.0*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A1560")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3ECF7")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#702080")),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (0, -1), (0, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8CCE2")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    activos = [r for r in filas if r["pedidos"] >= 1]
    inactivos = [r for r in filas if r["pedidos"] == 0]

    # 1) Tabla completa (la original)
    elementos += [banda(f"TODOS LOS ASOCIADOS  ({len(filas)})", "#702080"), Spacer(1, 4),
                  construir_tabla(filas)]
    # 2) Activas
    elementos += [Spacer(1, 18), banda(f"ASOCIADOS ACTIVOS  ({len(activos)})", "#F25C00"),
                  Spacer(1, 4), construir_tabla(activos)]
    # 3) Inactivas
    elementos += [Spacer(1, 18), banda(f"ASOCIADOS INACTIVOS  ({len(inactivos)})", "#8A7E93"),
                  Spacer(1, 4), construir_tabla(inactivos)]

    doc.build(elementos)
    buf.seek(0)
    return buf


# ============================================================================
# GOOGLE SHEETS (gratis; requiere credentials.json de Google Cloud junto a app.py)
# Crea una hoja nueva en TU Google Drive con los datos y devuelve su URL.
# La primera vez abre el navegador para que autorices; luego guarda token.json
# y ya no vuelve a pedir permiso.
# ============================================================================
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",  # solo archivos creados por esta app
]


def _ruta_junto_app(nombre):
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), nombre)


def _leer_secret_json(nombre):
    """Lee un JSON guardado en Streamlit Secrets como texto multilínea."""
    try:
        valor = st.secrets.get(nombre, "")
        if valor:
            return json.loads(str(valor))
    except Exception:
        pass
    return None


def _credenciales_google():
    """Obtiene credenciales OAuth para Google Sheets.

    En local puede usar credentials.json/token.json.
    En Streamlit Cloud usa GOOGLE_TOKEN_JSON y GOOGLE_CREDENTIALS_JSON desde Secrets.
    """
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow

    cred_path = _ruta_junto_app("credentials.json")
    token_path = _ruta_junto_app("token.json")

    creds = None

    # 1) NUBE: token guardado en Streamlit Secrets.
    token_info = _leer_secret_json("GOOGLE_TOKEN_JSON")
    if token_info:
        creds = Credentials.from_authorized_user_info(token_info, GOOGLE_SCOPES)

    # 2) LOCAL: token.json junto a app.py.
    elif os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, GOOGLE_SCOPES)

    # 3) Refrescar token si está vencido.
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())

    # 4) Si ya es válido, listo.
    if creds and creds.valid:
        return creds

    # 5) Si no hay token válido, solo local permite abrir navegador para autorizar.
    cred_info = _leer_secret_json("GOOGLE_CREDENTIALS_JSON")
    if cred_info:
        raise RuntimeError(
            "Google Sheets está configurado con credentials en Secrets, pero falta "
            "GOOGLE_TOKEN_JSON válido. Genera token.json en tu laptop y pégalo en Secrets."
        )

    if not os.path.exists(cred_path):
        raise RuntimeError(
            "Falta 'credentials.json' junto a app.py o GOOGLE_CREDENTIALS_JSON/GOOGLE_TOKEN_JSON "
            "en Streamlit Secrets."
        )

    flow = InstalledAppFlow.from_client_secrets_file(cred_path, GOOGLE_SCOPES)
    creds = flow.run_local_server(port=0)  # abre el navegador para autorizar en local
    with open(token_path, "w", encoding="utf-8") as f:
        f.write(creds.to_json())
    return creds


def subir_a_google_sheets(filas, titulo):
    """Crea una hoja de cálculo en el Drive del usuario con 3 pestañas y devuelve su URL."""
    import gspread
    creds = _credenciales_google()
    gc = gspread.authorize(creds)
    sh = gc.create(titulo.strip() or "Reporte de Asociados")

    MORADO = {"red": 112/255, "green": 32/255, "blue": 128/255}   # #702080
    LILA   = {"red": 243/255, "green": 236/255, "blue": 247/255}  # #F3ECF7
    BLANCO = {"red": 1, "green": 1, "blue": 1}

    def llenar(ws, registros):
        total_ped = sum(r["pedidos"] for r in registros)
        total_ing = sum(r["ingresos"] for r in registros)
        data = [["Nombre del Asociado", "Total de Pedidos", "Total Compras Catálogo Actual"]]
        for r in registros:
            data.append([r["nombre"], r["pedidos"], r["ingresos"]])
        data.append(["TOTAL", total_ped, total_ing])
        sh.values_update(f"'{ws.title}'!A1",
                         params={"valueInputOption": "USER_ENTERED"},
                         body={"values": data})
        n = len(data)
        try:  # el formato es opcional: si falla, los datos igual quedan bien
            enc = {"backgroundColor": MORADO,
                   "textFormat": {"bold": True, "foregroundColor": BLANCO},
                   "horizontalAlignment": "CENTER"}
            ws.format("A1:C1", enc)                    # encabezado morado
            ws.format(f"A{n}:C{n}", enc)               # fila TOTAL morada
            ws.format(f"C2:C{n}", {"numberFormat": {"type": "CURRENCY", "pattern": '"$"#,##0'}})
            ws.freeze(rows=1)
            reqs = [
                {"updateDimensionProperties": {
                    "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
                    "properties": {"pixelSize": 300}, "fields": "pixelSize"}},
                {"updateDimensionProperties": {
                    "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 3},
                    "properties": {"pixelSize": 140}, "fields": "pixelSize"}},
            ]
            if n > 3:  # bandas alternadas (zebra) nativas en las filas de datos: 1 sola petición
                reqs.append({"addBanding": {"bandedRange": {
                    "range": {"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n - 1,
                              "startColumnIndex": 0, "endColumnIndex": 3},
                    "rowProperties": {"firstBandColor": BLANCO, "secondBandColor": LILA}}}})
            sh.batch_update({"requests": reqs})
        except Exception:
            pass

    activos = [r for r in filas if r["pedidos"] >= 1]
    inactivos = [r for r in filas if r["pedidos"] == 0]

    ws1 = sh.sheet1
    ws1.update_title("Reporte")
    llenar(ws1, filas)                                                    # tabla completa
    llenar(sh.add_worksheet(title="Activos", rows=max(len(activos) + 6, 12), cols=4), activos)
    llenar(sh.add_worksheet(title="Inactivos", rows=max(len(inactivos) + 6, 12), cols=4), inactivos)
    return sh.url


# ============================================================================
# ACCESO PRIVADO
# ============================================================================
def _obtener_clave_app():
    """Obtiene la contraseña de acceso desde Secrets o variable de entorno."""
    try:
        clave = st.secrets.get("CLAVE_APP", "")
        if clave:
            return str(clave)
    except Exception:
        pass
    return os.environ.get("CLAVE_APP", "")


def _validar_acceso():
    """Muestra pantalla de contraseña y devuelve True si el usuario ya entró."""
    import hmac

    clave_correcta = _obtener_clave_app().strip()

    # Si no configuras CLAVE_APP, la app queda abierta. Útil para pruebas locales.
    if not clave_correcta:
        return True

    if st.session_state.get("acceso_autorizado"):
        with st.sidebar:
            st.success("🔒 Sesión iniciada")
            if st.button("Cerrar sesión", use_container_width=True):
                st.session_state.pop("acceso_autorizado", None)
                st.rerun()
        return True

    # Estilos solo para la pantalla de acceso.
    # También ocultamos el texto de Streamlit “Press Enter to apply”, que hacía
    # que el campo se viera amontonado junto al ícono de mostrar contraseña.
    st.markdown("""
        <style>
        div[data-testid="InputInstructions"] { display: none !important; }

        .login-wrap {
            max-width: 520px;
            margin: 7vh auto 0 auto;
            padding: 0 14px;
        }
        .login-card {
            background: #ffffff;
            border: 1px solid #eadff0;
            border-radius: 24px;
            box-shadow: 0 18px 45px rgba(112, 32, 128, .13);
            padding: 34px 34px 30px 34px;
            text-align: center;
        }
        .login-icon {
            width: 64px;
            height: 64px;
            margin: 0 auto 14px auto;
            border-radius: 20px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: linear-gradient(135deg, #702080, #F25C00);
            color: white;
            font-size: 30px;
            box-shadow: 0 12px 25px rgba(112, 32, 128, .22);
        }
        .login-title {
            margin: 0;
            color: #702080;
            font-size: 28px;
            font-weight: 800;
            line-height: 1.1;
        }
        .login-subtitle {
            margin: 12px 0 22px 0;
            color: #66586c;
            font-size: 15px;
            line-height: 1.45;
        }
        .login-label {
            text-align: left;
            font-weight: 700;
            font-size: 14px;
            color: #4a1560;
            margin: 2px 0 8px 2px;
        }

        /* Campo de contraseña */
        .stTextInput > div > div {
            border-radius: 14px !important;
        }
        .stTextInput input {
            height: 48px !important;
            border-radius: 14px !important;
            font-size: 16px !important;
        }

        /* Botón de entrar */
        div[data-testid="stButton"] button {
            background: #702080 !important;
            color: #ffffff !important;
            font-weight: 800 !important;
            border-radius: 14px !important;
            border: none !important;
            height: 48px !important;
            margin-top: 8px !important;
            box-shadow: 0 10px 22px rgba(112, 32, 128, .18);
        }
        div[data-testid="stButton"] button:hover {
            background: #5a1a6b !important;
        }
        div[data-testid="stButton"] button * {
            color: #ffffff !important;
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <div class="login-wrap">
            <div class="login-card">
                <div class="login-icon">🔐</div>
                <h2 class="login-title">Acceso privado</h2>
                <p class="login-subtitle">
                    Ingresa la contraseña para usar el extractor de reportes Betterware.
                </p>
    """, unsafe_allow_html=True)

    # Pantalla de acceso sin st.form para evitar el aviso de Streamlit:
    # “Missing Submit Button”. El botón valida la contraseña.
    st.markdown('<div class="login-label">Contraseña de acceso</div>', unsafe_allow_html=True)
    clave_ingresada = st.text_input(
        "Contraseña de acceso",
        type="password",
        placeholder="Escribe la contraseña",
        label_visibility="collapsed",
        key="clave_login"
    )
    entrar = st.button("Entrar", use_container_width=True)

    st.markdown("""
            </div>
        </div>
    """, unsafe_allow_html=True)

    if entrar:
        if hmac.compare_digest(clave_ingresada.strip(), clave_correcta):
            st.session_state["acceso_autorizado"] = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta. Revisa e intenta de nuevo.")

    st.stop()


# ============================================================================
# INTERFAZ
# ============================================================================
def main():
    st.set_page_config(page_title="Reporte de Asociados", page_icon="📊", layout="centered")
    st.markdown("""
        <style>
        .stApp { background-color: #faf6fb !important; }

        h1 { color: #702080 !important; }
        h2, h3, h5 { color: #4a1560 !important; }
        .stCaption, .stMarkdown p { color: #2b2330 !important; }

        /* Oculta instrucciones pequeñas como “Press Enter to apply” */
        div[data-testid="InputInstructions"] { display: none !important; }

        /* Inputs generales de la app */
        .stTextInput input {
            background-color:#fff !important;
            color:#2b2330 !important;
            border:1px solid #702080 !important;
            border-radius: 10px !important;
        }

        /* Uploader */
        [data-testid="stFileUploader"] {
            background:#fff !important;
            border:1px dashed #702080 !important;
            border-radius:12px;
        }
        [data-testid="stFileUploaderDropzone"],
        [data-testid="stFileUploaderDropzone"] section,
        section[data-testid="stFileUploaderDropzone"] {
            background:#f4ecf7 !important;
            color:#2b2330 !important;
        }
        [data-testid="stFileUploaderDropzone"] button,
        [data-testid="stBaseButton-secondary"] {
            background:#702080 !important;
            color:#fff !important;
            border:none !important;
        }
        [data-testid="stFileUploaderDropzone"] button *,
        [data-testid="stBaseButton-secondary"] * {
            color:#fff !important;
        }

        /* Botones */
        .stButton>button,
        .stDownloadButton>button {
            background:#702080 !important;
            color:#fff !important;
            font-weight:bold;
            border-radius:10px;
            border:none;
            padding:10px 20px;
        }
        .stButton>button:hover,
        .stDownloadButton>button:hover {
            background:#5a1a6b !important;
        }
        .stButton>button *,
        .stDownloadButton>button * {
            color:#fff !important;
        }
        </style>
    """, unsafe_allow_html=True)

    _validar_acceso()

    # Encabezado de marca: logo + nombre INNquietus + título (logo embebido en base64
    # para que quede alineado junto al texto; usa 'logo.png' junto a app.py si existe).
    import base64
    logo_path = _ruta_junto_app("logo.png")
    logo_html = ""
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as _f:
            _b64 = base64.b64encode(_f.read()).decode()
        logo_html = f"<img src='data:image/png;base64,{_b64}' style='height:88px;width:auto;'>"

    st.markdown(f"""
        <div style="display:flex;align-items:center;gap:18px;margin-bottom:4px;">
          {logo_html}
          <div style="line-height:1.05;">
            <div style="font-size:14px;font-weight:800;letter-spacing:3px;
                        text-transform:uppercase;color:#F25C00;">INN<span style="color:#702080;">quietus</span></div>
            <h1 style="color:#702080;margin:2px 0 0;font-size:42px;">Reporte de Asociados</h1>
          </div>
        </div>
        <div style="height:6px;border-radius:3px;margin:8px 0 10px;
             background:linear-gradient(90deg,#702080 0 33%,#F2E500 33% 66%,#F25C00 66% 100%);"></div>
    """, unsafe_allow_html=True)
    st.caption("Herramienta privada para procesar reportes CTBT Betterware, generar Excel/PDF y crear Google Sheets.")

    titulo_reporte = st.text_input("Título del reporte", value="Reporte de Asociados")

    st.markdown("##### 🤖 Corregir nombres con IA (recomendado)")
    usar_ia = st.checkbox("Usar IA para reconstruir los nombres", value=True)
    proveedor, api_key = "Claude", ""
    if usar_ia:
        proveedor = st.radio("Proveedor de IA", ["Claude", "Gemini"], horizontal=True, index=0)
        if proveedor == "Claude":
            env_var = "ANTHROPIC_API_KEY"
        elif proveedor == "Grok":
            env_var = "XAI_API_KEY"
        else:
            env_var = "GEMINI_API_KEY"
        env_key = os.environ.get(env_var, "")
        if env_key:
            # La key viene del .env: NO se muestra en pantalla por seguridad.
            api_key = env_key
            st.caption(f"🔒 API Key de {proveedor} cargada del archivo .env (oculta por seguridad).")
        else:
            # No hay .env: se pide, pero el campo es de tipo password (puntos).
            api_key = st.text_input(
                f"API Key de {proveedor}", value="", type="password",
                help="Se usa solo en esta sesión. No se guarda en el código.")

    archivo = st.file_uploader("Arrastra o selecciona el PDF", type=["pdf"])

    # ========================================================================
    # PROCESO (esto SÍ consume créditos de IA): corre SOLO al pulsar el botón.
    # El resultado se guarda en st.session_state para no repetirlo en cada
    # recarga de la interfaz. Así, tocar otros controles o descargar archivos
    # NO vuelve a llamar a la IA ni gasta créditos. Solo se renueva cuando
    # vuelves a pulsar "Procesar PDF" con un archivo.
    # ========================================================================
    if archivo is not None and st.button("🚀 Procesar PDF"):
        with st.spinner("Leyendo el PDF…"):
            pdf_bytes = archivo.read()
            try:
                filas, texto_crudo = extraer(pdf_bytes)
            except Exception as e:
                st.error(f"Error al leer el PDF: {e}"); st.stop()
        if not filas:
            st.error("No detecté filas de asociados. ¿El PDF tiene el formato esperado?"); st.stop()

        aviso_ia = None  # (tipo, mensaje[, detalle]) para mostrarlo luego desde la sesión
        if usar_ia and api_key.strip():
            with st.spinner(f"Corrigiendo nombres con {proveedor}…"):
                try:
                    filas = limpiar_nombres_con_ia(texto_crudo, filas, proveedor, api_key.strip())
                    aviso_ia = ("success", "Nombres corregidos con IA ✅")
                except Exception as e:
                    import traceback
                    aviso_ia = ("error", f"La IA falló ({e}). Muestro los nombres sin corregir.",
                                traceback.format_exc())
        elif usar_ia:
            aviso_ia = ("warning", "Marcaste usar IA pero falta la API Key. Procesé sin IA.")

        filas.sort(key=lambda r: (-r["ingresos"], -r["pedidos"]))

        total_extraido = sum(r["ingresos"] for r in filas)
        pedidos_extraidos = sum(r["pedidos"] for r in filas)
        total_oficial = extraer_total_oficial_catalogo_actual(texto_crudo)
        aviso_total = None
        if total_oficial:
            pedidos_oficiales, compras_oficiales, linea_total = total_oficial
            if compras_oficiales != total_extraido or pedidos_oficiales != pedidos_extraidos:
                aviso_total = (
                    pedidos_oficiales,
                    compras_oficiales,
                    pedidos_extraidos,
                    total_extraido,
                    linea_total,
                )

        # Se genera el Excel y el PDF una sola vez y se guardan como bytes en la sesión.
        st.session_state["resultado"] = {
            "filas": filas,
            "titulo": titulo_reporte,
            "excel": generar_excel(filas, titulo_reporte).getvalue(),
            "pdf": generar_pdf(filas, titulo_reporte).getvalue(),
            "aviso_ia": aviso_ia,
            "aviso_total": aviso_total,
        }
        st.session_state.pop("sheets_url", None)  # el link de Sheets es de otro reporte

    # ========================================================================
    # RENDER (gratis): se dibuja en CADA recarga leyendo de la sesión.
    # No reprocesa nada; por eso el resultado permanece en pantalla aunque
    # toques la interfaz, y solo cambia al procesar otro PDF.
    # ========================================================================
    res = st.session_state.get("resultado")
    if res:
        aviso = res.get("aviso_ia")
        if aviso:
            if aviso[0] == "success":
                st.success(aviso[1])
            elif aviso[0] == "warning":
                st.warning(aviso[1])
            elif aviso[0] == "error":
                st.error(aviso[1])
                if len(aviso) > 2:
                    with st.expander("Ver detalle técnico del error"):
                        st.code(aviso[2])

        filas = res["filas"]
        aviso_total = res.get("aviso_total")
        if aviso_total:
            pedidos_oficiales, compras_oficiales, pedidos_extraidos, total_extraido, linea_total = aviso_total
            diferencia = total_extraido - compras_oficiales
            signo = "+" if diferencia > 0 else ""
            st.warning(
                f"⚠️ Diferencia de total: PDF ${compras_oficiales:,} vs suma ${total_extraido:,} "
                f"({signo}${diferencia:,}). Se conserva la suma por asociada."
            )

        st.success(f"✅ Procesado: {len(filas)} asociados. "
                   "El resultado se queda en pantalla; solo se renueva al procesar otro PDF.")
        st.dataframe(
            [{"Nombre del Asociado": r["nombre"], "Total de Pedidos": r["pedidos"],
              "Total Compras Catálogo Actual": f"${r['ingresos']:,}"} for r in filas],
            width="stretch", hide_index=True)

        col1, col2 = st.columns(2)
        with col1:
            st.download_button("⬇️ Descargar Excel", res["excel"], file_name="Reporte_Asociados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col2:
            st.download_button("⬇️ Descargar PDF", res["pdf"], file_name="Reporte_Asociados.pdf",
                mime="application/pdf")

        # --- Abrir en Google Sheets (usa el resultado YA procesado; NO llama a la IA) ---
        st.markdown("##### 📗 Abrir en Google Sheets")
        if st.button("Crear y abrir en Google Sheets"):
            with st.spinner("Creando la hoja en tu Google Drive…"):
                try:
                    url = subir_a_google_sheets(res["filas"], res["titulo"])
                    st.session_state["sheets_url"] = url
                    try:
                        import webbrowser
                        webbrowser.open_new_tab(url)  # abre en el navegador de tu equipo
                    except Exception:
                        pass
                    st.success("Hoja creada ✅")
                except ModuleNotFoundError:
                    st.error(
                        "Google Sheets no está disponible porque faltan dependencias. "
                        "Revisa que requirements.txt esté actualizado y reinicia/despliega de nuevo la app."
                    )
                except Exception as e:
                    st.error(f"No se pudo crear la hoja: {e}")
                    import traceback
                    with st.expander("Ver detalle técnico"):
                        st.code(traceback.format_exc())
        if st.session_state.get("sheets_url"):
            st.markdown(f"🔗 [Abrir la hoja en Google Sheets]({st.session_state['sheets_url']})")

    # Pie de página de marca
    st.markdown(
        "<hr style='border:none;border-top:1px solid #e7dcef;margin:36px 0 10px;'>"
        "<div style='text-align:center;color:#9b8aa6;font-size:12px;'>"
        "<b style='color:#702080;'>INNquietus</b> · Reporte de Asociados</div>",
        unsafe_allow_html=True)


if __name__ == "__main__":
    main()